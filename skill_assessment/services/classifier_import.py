# route: (service) | file: skill_assessment/services/classifier_import.py
"""Импорт таксономии из Excel классификатора (лист «Классификатор_навыков»)."""

from __future__ import annotations

import hashlib
import io
import uuid
from typing import Any

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from skill_assessment.infrastructure.db_models import SkillDomainRow, SkillRow
from skill_assessment.schemas.api import ClassifierImportOut


def _domain_code(title: str) -> str:
    h = hashlib.md5(title.strip().encode("utf-8")).hexdigest()[:10]
    return f"D_{h}"


def import_classifier_xlsx(db: Session, xlsx_bytes: bytes) -> ClassifierImportOut:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise HTTPException(
            status_code=503,
            detail="openpyxl_required: pip install openpyxl",
        ) from e

    if not xlsx_bytes:
        raise HTTPException(status_code=400, detail="empty_file")

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        name = "Классификатор_навыков"
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="classifier_sheet_empty")

        header = [_norm_header(c) for c in header_row]
        col = _build_col_index(header)

        required = ("skill_id", "domain", "skill_name")
        for r in required:
            if r not in col:
                raise HTTPException(
                    status_code=400,
                    detail=f"classifier_missing_column:{r}",
                )

        domains_created = 0
        skills_created = 0
        skills_updated = 0
        domain_cache: dict[str, SkillDomainRow] = {}

        for raw in rows_iter:
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            skill_id_cell = _cell(raw, col["skill_id"])
            domain_title = _cell(raw, col["domain"])
            skill_name = _cell(raw, col["skill_name"])
            if not skill_id_cell or not domain_title or not skill_name:
                continue

            code = _short_code(skill_id_cell)
            if domain_title not in domain_cache:
                row_d = db.scalars(
                    select(SkillDomainRow).where(SkillDomainRow.title == domain_title)
                ).first()
                if row_d is None:
                    row_d = SkillDomainRow(
                        id=str(uuid.uuid4()),
                        code=_domain_code(domain_title),
                        title=domain_title,
                    )
                    db.add(row_d)
                    domains_created += 1
                domain_cache[domain_title] = row_d

            dom = domain_cache[domain_title]
            existing = db.scalars(select(SkillRow).where(SkillRow.code == code)).first()
            if existing:
                changed = False
                if existing.title != skill_name:
                    existing.title = skill_name
                    changed = True
                if existing.domain_id != dom.id:
                    existing.domain_id = dom.id
                    changed = True
                if changed:
                    skills_updated += 1
            else:
                db.add(
                    SkillRow(
                        id=str(uuid.uuid4()),
                        domain_id=dom.id,
                        code=code,
                        title=skill_name,
                    )
                )
                skills_created += 1

        db.commit()
        return ClassifierImportOut(
            sheet_used=ws.title,
            domains_created=domains_created,
            skills_created=skills_created,
            skills_updated=skills_updated,
        )
    finally:
        wb.close()


def _norm_header(c: Any) -> str:
    if c is None:
        return ""
    return str(c).strip().lower()


def _build_col_index(header: list[str]) -> dict[str, int]:
    aliases = {
        "skill_id": {"skill_id", "id"},
        "department": {"department", "подразделение", "dept"},
        "domain": {"domain", "домен", "cluster", "кластер"},
        "skill_name": {"skill_name", "skill", "навык", "name"},
        "source": {"source", "источник"},
        "note_regulation_kpi": {"note_regulation_kpi", "note", "комментарий"},
    }
    idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        for canon, variants in aliases.items():
            if h in variants and canon not in idx:
                idx[canon] = i
                break
    return idx


def _cell(row: tuple[Any, ...], i: int) -> str:
    if i >= len(row):
        return ""
    v = row[i]
    if v is None:
        return ""
    return str(v).strip()


def _short_code(skill_id: str) -> str:
    s = skill_id.strip()
    if len(s) > 64:
        return s[:64]
    return s
