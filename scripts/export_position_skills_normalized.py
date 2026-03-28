# -*- coding: utf-8 -*-
"""
Нормализованная таблица для БД: одна строка = одна связь должность ↔ навык.

Колонки: id, position_code, position_name_ru, skill_rank (1…7), skill_name.

Вывод в repo/exports/: position_skills_normalized.xlsx и .csv (UTF-8 с BOM для Excel).

Запуск из корня пакета skill_assessment:
    python scripts/export_position_skills_normalized.py
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from skill_assessment.data.top20_position_skills import TOP20_POSITION_SKILL_ROWS

# repo/exports — на уровень выше каталога пакета skill_assessment (pyproject)
REPO_EXPORTS = Path(__file__).resolve().parents[2] / "exports"


def iter_normalized_rows():
    """Генерация плоских строк (id, code, name_ru, rank, skill_name)."""
    n = 0
    for code, title, skills in TOP20_POSITION_SKILL_ROWS:
        for rank, skill in enumerate(skills, start=1):
            n += 1
            yield (n, code, title, rank, skill.strip())


def main() -> None:
    REPO_EXPORTS.mkdir(parents=True, exist_ok=True)
    rows = list(iter_normalized_rows())

    headers = ["id", "position_code", "position_name_ru", "skill_rank", "skill_name"]

    xlsx_path = REPO_EXPORTS / "position_skills_normalized.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "position_skills"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 56
    wb.save(xlsx_path)

    csv_path = REPO_EXPORTS / "position_skills_normalized.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        w.writerow(headers)
        w.writerows(rows)

    print(f"Written: {xlsx_path}")
    print(f"Written: {csv_path}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
