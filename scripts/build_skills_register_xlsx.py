# route: (script) | file: skill_assessment/scripts/build_skills_register_xlsx.py
"""Сборка Excel-реестра навыков из матрицы Description + справочник заказчика (PDF)."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DESC = ROOT.parent / "Description"
OUT = DESC / "skills_register_matrix.xlsx"
MATRIX_MD = DESC / "matritsa_skills.md"

# Из PDF «Некоторые соображения по архитектуре» — 4 домена × 9 навыков (англ. ярлыки в скобках по PDF).
CUSTOMER_4x9 = {
    "Коммуникации": [
        ("Переговоры", "Negotiation"),
        ("Презентация", "Presentation"),
        ("Убеждение", "Persuasion"),
        ("Активное слушание", "Active Listening"),
        ("Разрешение конфликтов", "Conflict Resolution"),
        ("Рассказывание историй", "Storytelling"),
        ("Нетворкинг", "Networking"),
        ("Обратная связь", "Feedback"),
        ("Публичные выступления", "Public Speaking"),
    ],
    "Лидерство": [
        ("Принятие решений", "Decision Making"),
        ("Мотивация команды", "Team Motivation"),
        ("Делегирование", "Delegation"),
        ("Стратегическое мышление", "Strategic Thinking"),
        ("Ответственность", "Responsibility"),
        ("Управление изменениями", "Change Management"),
        ("Влияние", "Influence"),
        ("Коучинг", "Coaching"),
        ("Формирование видения", "Vision Building"),
    ],
    "Системность": [
        ("Аналитическое мышление", "Analytical Thinking"),
        ("Проектирование процессов", "Process Design"),
        ("Планирование", "Planning"),
        ("Оценка рисков", "Risk Assessment"),
        ("Интерпретация данных", "Data Interpretation"),
        ("Оптимизация", "Optimization"),
        ("Приоритизация", "Prioritization"),
        ("Управление проектами", "Project Management"),
        ("Документация", "Documentation"),
    ],
    "Творчество": [
        ("Генерация идей", "Idea Generation"),
        ("Формулирование проблемы", "Problem Framing"),
        ("Разработка эксперимента", "Experiment Design"),
        ("Инновационное мышление", "Innovation Thinking"),
        ("Сценарное моделирование", "Scenario Modeling"),
        ("Креативная стратегия", "Creative Strategy"),
        ("Межотраслевое мышление", "Cross-domain Thinking"),
        ("Построение гипотез", "Hypothesis Building"),
        ("Быстрое прототипирование", "Rapid Prototyping"),
    ],
}


def parse_matrix_md(path: Path) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = []
    text = path.read_text(encoding="utf-8")
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|") or line.startswith("|---"):
            continue
        parts = [p.strip() for p in line.split("|")]
        parts = [p for p in parts if p]
        if len(parts) < 3:
            continue
        if parts[0].lower() in ("функция", "function"):
            continue
        func, domain, skills_blob = parts[0], parts[1], parts[2]
        for raw in skills_blob.split(";"):
            name = raw.strip()
            if name:
                rows.append((func, domain, name))
    return rows


def sheet_matrix(wb: Workbook, rows: list[tuple[str, str, str]]) -> None:
    ws = wb.create_sheet("Матрица_6функций_6доменов", 0)
    headers = [
        "skill_id",
        "function_block",
        "domain",
        "skill_name",
        "source_file",
        "binding_hint",
    ]
    ws.append(headers)
    for i, (func, domain, skill) in enumerate(rows, start=1):
        sid = f"SK_{i:04d}"
        hint = (
            "Привязка: выберите подразделения/должности в профиле организации; "
            "роль — через матрицу компетенций должности."
        )
        ws.append([sid, func, domain, skill, "Description/matritsa_skills.md", hint])
    _style_header(ws)
    _autosize(ws)


def sheet_customer(wb: Workbook) -> None:
    ws = wb.create_sheet("Заказчик_4домена_x9", 1)
    ws.append(["skill_id", "customer_domain", "skill_name_ru", "label_en", "source", "binding_hint"])
    n = 1
    for domain, items in CUSTOMER_4x9.items():
        for ru, en in items:
            ws.append(
                [
                    f"CUST_{n:04d}",
                    domain,
                    ru,
                    en,
                    "Description: PDF «Некоторые соображения по архитектуре»",
                    "Согласовать с HR: какой функциональный блок и должности используют эту линию.",
                ]
            )
            n += 1
    _style_header(ws)
    _autosize(ws)


def sheet_recommendations(wb: Workbook) -> None:
    ws = wb.create_sheet("Рекомендации_по_реестру", 2)
    ws.column_dimensions["A"].width = 100
    recs = [
        (
            "1. Что хранить в системе",
            "Единый справочник навыков (Skill) + домены (SkillDomain). Индикаторы и уровни — второй этап (служебная записка: 3–5 индикаторов на skill).",
        ),
        (
            "2. Куда «вешать» навыки",
            "Логичная цепочка: Организация → подразделение (org unit) → должность/позиция → профиль навыков (набор skill + вес/обязательность). Роль в IAM (доступ) и «функция» в HR (HR/Маркетинг/…) — разные оси: не смешивать в одной таблице без явной связи.",
        ),
        (
            "3. Две линии таксономии",
            "Матрица 6 функций × 6 доменов (matritsa_skills) — для корпоративного каркаса. Линия заказчика 4×9 из PDF — для пилота/сверки; сопоставление через код или отдельное поле «альтернативное имя».",
        ),
        (
            "4. MVP",
            "Сначала зафиксировать реестр и 1–2 навыка на пилотное подразделение (как в вашем PDF), затем расширять. Микроскиллы в оценке — позже.",
        ),
        (
            "5. Excel и система",
            "Этот файл — мастер-данные для согласования с заказчиком; загрузка в БД — отдельный импорт или ручной ввод в админке.",
        ),
    ]
    for title, body in recs:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([body])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.append([])


def _style_header(ws) -> None:
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)


def _autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row + 1, 500)):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)


def main() -> None:
    if not MATRIX_MD.is_file():
        raise SystemExit(f"Не найден файл: {MATRIX_MD}")
    rows = parse_matrix_md(MATRIX_MD)
    wb = Workbook()
    wb.remove(wb.active)
    sheet_matrix(wb, rows)
    sheet_customer(wb)
    sheet_recommendations(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Written: {OUT} ({len(rows)} skills in matrix + customer sheet)")


if __name__ == "__main__":
    main()
