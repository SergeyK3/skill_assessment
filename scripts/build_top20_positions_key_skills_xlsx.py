# -*- coding: utf-8 -*-
"""
Таблица: 20 типовых должностей × 5–7 ключевых навыков (широкий формат: навыки в колонках).

Запуск из корня пакета skill_assessment: python scripts/build_top20_positions_key_skills_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from skill_assessment.data.top20_position_skills import TOP20_POSITION_SKILL_ROWS

OUT = Path(__file__).resolve().parents[1] / "docs" / "top20_positions_key_skills.xlsx"


def main() -> None:
    ROWS = TOP20_POSITION_SKILL_ROWS
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Должности_и_навыки"

    headers = [
        "№",
        "position_code",
        "Должность (RU)",
        "Кол-во навыков",
        "Навык 1",
        "Навык 2",
        "Навык 3",
        "Навык 4",
        "Навык 5",
        "Навык 6",
        "Навык 7",
    ]
    ws.append(headers)
    for i, (code, title, skills) in enumerate(ROWS, start=1):
        pad = skills + [""] * (7 - len(skills))
        row = [i, code, title, len(skills)] + pad[:7]
        ws.append(row)

    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18 if c <= 4 else 36

    wb.save(OUT)
    print(f"Written: {OUT}")


if __name__ == "__main__":
    main()
